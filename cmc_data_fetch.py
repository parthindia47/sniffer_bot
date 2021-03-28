'''
@author: parth pandya

about :
    this file creates an excel sheet which provides all coin information
    like twitter account , website etc.


development :
    it seems sanbox feature in coin market cap in not proper.
    earlier there were two keys , one for sandbox and other for pro.
    it seems now there is only one key.

    key is gained by account :
        ravi.kumar.engineer.2000@gmail.com

    rate limit per day is : 330 calls
    per minut is : 30 calls
    so it is necessary that we sleep for 3 seconds.
'''

from requests import Request, Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
import json
import csv
import math
import openpyxl
from datetime import datetime
import time as sleep

url_map  = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/map'
url_info = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/info'

parameters_map = {
}

parameters_info_ids = {
  'id': '1,2'
}

headers = {
  'Accepts': 'application/json',
  'X-CMC_PRO_API_KEY': '7f45549a-0549-4f55-8763-d15a5631056f',
}


Map_details_list = [  'Dummy',
                      'Main_id' ,
                      'Name',
                      'Symbol',
                      'URL_name',
                      'Is_Active',
                      'status',
                      'first_historical_data',
                      'last_historical_data',
                      'platform',
                      'platform_name',
                      'platform_token_address' ,
                      'catagory_coin_tocken',
                      'logo_url',
                      'description',
                      'date_added',
                      'notice',
                      'tags_minable_or_not',
                      'website',
                      'technical_doc',
                      'explorer',
                      'source_code',
                      'message_board',
                      'chat',
                      'announcement',
                      'reddit',
                      'twitter'
                    ]


session = Session()
session.headers.update(headers)

#creating an excel file for writing data.
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "CMC_details"

# writing headers in excel file
# carefull about index
for i in range(1, len(Map_details_list)):
  c1 = sheet.cell(row = 1, column = i)
  c1.value = Map_details_list[i]


#fetching list of all availible currency and ids.
try:
  response = session.get(url_map, params=parameters_map)
  data_map = json.loads(response.text)
  data_map_list = data_map['data']
  length_map = len(data_map_list)
  id_list = []

  for i in range(length_map):

    id_list.append(data_map_list[i]['id'])

    sheet.cell(row = i+2, column = Map_details_list.index('Main_id')).value = data_map_list[i]['id']
    sheet.cell(row = i+2, column = Map_details_list.index('Name')).value = data_map_list[i]['name']
    sheet.cell(row = i+2, column = Map_details_list.index('Symbol')).value = data_map_list[i]['symbol']
    sheet.cell(row = i+2, column = Map_details_list.index('URL_name')).value = data_map_list[i]['slug']
    sheet.cell(row = i+2, column = Map_details_list.index('Is_Active')).value = data_map_list[i]['is_active']
    sheet.cell(row = i+2, column = Map_details_list.index('status')).value = "NA"

    if 'first_historical_data' in data_map_list[i] :
        sheet.cell(row = i+2, column = Map_details_list.index('first_historical_data')).value = \
        data_map_list[i]['first_historical_data']
    else :
        sheet.cell(row = i+2, column = Map_details_list.index('first_historical_data')).value = "NA"

    if 'last_historical_data' in data_map_list[i] :
        sheet.cell(row = i+2, column = Map_details_list.index('last_historical_data')).value = \
        data_map_list[i]['last_historical_data']
    else :
        sheet.cell(row = i+2, column = Map_details_list.index('last_historical_data')).value = "NA"

    if data_map_list[i]['platform']:

      curr_platform_dict = data_map_list[i]['platform']

      sheet.cell(row = i+2, column = Map_details_list.index('platform')).value = "YES"
      sheet.cell(row = i+2, column = Map_details_list.index('platform_name')).value = \
      curr_platform_dict['name']
      sheet.cell(row = i+2, column = Map_details_list.index('platform_token_address')).value = \
      curr_platform_dict['token_address']

    else:

      sheet.cell(row = i+2, column = Map_details_list.index('platform')).value = "NA"

except (ConnectionError, Timeout, TooManyRedirects) as e:
  print("There is some issue fetching ID map , here is the error !!! " )
  print(e)



#fetching list of twitter accounts details. we will shift in steps of 100.

length_id = len(id_list)

for i in range(0, length_id, 100 ):

  sleep.sleep(3)

  print("getting max 100 coins from index: " + int(i) + "\n")

  Curr_id_list = list(id_list[i: i+100 if i+100 < length_id else length_id ])
  parameters_info_ids['id'] = ','.join(map(str, Curr_id_list))

  try:
    response = session.get(url_info, params=parameters_info_ids)
    data_info = json.loads(response.text)

    cnt = 0
    for j in range(i, i+100 if i+100 < length_id else length_id ):

      sheet.cell(row = j+2, column = Map_details_list.index('catagory_coin_tocken')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['category'])

      sheet.cell(row = j+2, column = Map_details_list.index('logo_url')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['logo'])

      sheet.cell(row = j+2, column = Map_details_list.index('description')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['description'])

      sheet.cell(row = j+2, column = Map_details_list.index('date_added')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['date_added'])

      sheet.cell(row = j+2, column = Map_details_list.index('notice')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['notice'])

      sheet.cell(row = j+2, column = Map_details_list.index('tags_minable_or_not')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['tags'])

      sheet.cell(row = j+2, column = Map_details_list.index('website')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['website'])

      sheet.cell(row = j+2, column = Map_details_list.index('technical_doc')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['technical_doc'])

      sheet.cell(row = j+2, column = Map_details_list.index('explorer')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['explorer'])

      sheet.cell(row = j+2, column = Map_details_list.index('source_code')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['source_code'])

      sheet.cell(row = j+2, column = Map_details_list.index('message_board')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['message_board'])

      sheet.cell(row = j+2, column = Map_details_list.index('chat')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['chat'])

      sheet.cell(row = j+2, column = Map_details_list.index('announcement')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['announcement'])

      sheet.cell(row = j+2, column = Map_details_list.index('reddit')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['reddit'])

      sheet.cell(row = j+2, column = Map_details_list.index('twitter')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['twitter'])

      cnt = cnt + 1

    #data_info = str(data_info).encode('utf8')
    #print(data_info)
    #print("************************************************************")
    #print("************************************************************")

  except (ConnectionError, Timeout, TooManyRedirects) as e:
    print("There is some issue fetching meta data of currency , here is the error !!! " )
    print(e)

wb.save("CMC_Coin_Details_Auto"+ "_" + str(datetime.now().date().day) + "_" + str(datetime.now().date().month) + ".xlsx")
print("file generation complete")
