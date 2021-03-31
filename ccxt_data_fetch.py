# -*- coding: utf-8 -*-
"""
Created on Sun Aug 16 16:43:36 2020

@author: I20035
"""

import ccxt
import time
import openpyxl


def get_specific_market( specific_market, all_market ):
    specific_market = "/" + specific_market
    mlen = len(specific_market)
    
    specific_market_list = []
    
    for market in all_market:
        if specific_market in market:
            specific_market_list.append(market[:-mlen])
            
    return specific_market_list

# column index start from 1.
def generate_specific_excel( excel_path , column_no, scan_list ):
    wb_obj = openpyxl.load_workbook( excel_path )
    sheet_obj = wb_obj.active
    
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    
    print("max col : " + str(max_col))
    print("max row : " + str(max_row))
    
    match_list = []
    
    # loop all row
    for i in range(2, max_row + 1): 
        cell_obj = sheet_obj.cell(row = i, column = column_no)
        if cell_obj.value in scan_list:
            match_list.append(cell_obj.value)
    
    return match_list
        

if __name__ == '__main__':
    
    exchange_list = ccxt.exchanges
    #print(ccxt.exchanges)
    
    binance  = ccxt.binance()
    binance_markets = binance.load_markets()
    
    symbols = binance.symbols                 # get a list of symbols
    binance_btc_markets_ccxt = get_specific_market( "BTC" , symbols )
    print( "ccxt len : " + str(len(binance_btc_markets_ccxt)))
    
#    markets_cmc = generate_specific_excel( "CMC_Coin_Details_Auto_16_8.xlsx" , 3 , binance_btc_markets_ccxt )
#    print( "cmc len : " + str(len(markets_cmc)))
    
    